"""
Offline IMU-only KNN prediction on pre-recorded data.

- Reads imu.csv files from Temporary_Data/Sample_* folders.
- Extracts time-domain and frequency-domain features.
- Applies a pre-trained scaler and KNN classifier.
- Logs predictions to an Excel file in Participant_Data.
- Pauses for 6 seconds whenever a stable Grab or Down is detected.

This script is intended as companion code for the associated publication.
"""

import os
import sys
import time
import re
from collections import Counter
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from scipy.fft import fft
from scipy.stats import entropy, kurtosis, skew

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

BASE_DIR = Path(__file__).resolve().parent

TEMPORARY_DATA_DIR = BASE_DIR / "Temporary_Data"
PARTICIPANT_DIR = BASE_DIR / "Participant_Data"
HEADER_CSV = BASE_DIR / "dataset_header.csv"
SCALER_PATH = BASE_DIR / "scaler.pkl"
KNN_PATH = BASE_DIR / "knn_model.pkl"

LABEL_MAP = {0: "Grab", 1: "Walk", 2: "Down"}

STOP_ALL = True
LINE_UP = "\033[1A"
LINE_CLEAR = "\x1b[2K"

# --------------------------------------------------------------------------- #
# Utilities
# --------------------------------------------------------------------------- #


def numerical_sort(value: str) -> int:
    """Extract the first integer from a string for natural sorting."""
    numbers = re.findall(r"\d+", value)
    return int(numbers[0]) if numbers else float("inf")


def extract_imu_features(df: pd.DataFrame) -> pd.Series:
    """
    Compute time-domain and frequency-domain features for each column
    of an IMU data frame.
    """
    feats = {}
    for col in df.columns:
        sig = df[col].values

        feats[f"{col}_mean"] = np.mean(sig)
        feats[f"{col}_std"] = np.std(sig)
        feats[f"{col}_min"] = np.min(sig)
        feats[f"{col}_max"] = np.max(sig)
        feats[f"{col}_range"] = np.ptp(sig)
        feats[f"{col}_median"] = np.median(sig)
        feats[f"{col}_skew"] = skew(sig)
        feats[f"{col}_kurtosis"] = kurtosis(sig)
        feats[f"{col}_energy"] = np.sum(sig ** 2)
        feats[f"{col}_rms"] = np.sqrt(np.mean(sig ** 2))

        hist, _ = np.histogram(sig, bins=20, density=True)
        feats[f"{col}_entropy"] = entropy(hist + 1e-6)

        fft_vals = np.abs(fft(sig))[: len(sig) // 2]  # type: ignore
        top_fft = np.sort(fft_vals)[-3:]
        for i, peak in enumerate(top_fft):
            feats[f"{col}_fft_peak_{i + 1}"] = peak

    return pd.Series(feats)


def init_participant_excel(participant_dir: Path) -> tuple[Workbook, object, Path]:
    """
    Create or open a participant Excel file in participant_dir.

    Returns
    -------
    wb : Workbook
        OpenPyXL workbook instance.
    ws : Worksheet
        Active worksheet.
    excel_path : Path
        Path to the Excel file.
    """
    participant_dir.mkdir(exist_ok=True)

    num = 1
    while True:
        excel_path = participant_dir / f"participant_IMU_{num}.xlsx"
        if not excel_path.exists():
            break
        num += 1

    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Participant_IMU{num}"  # type: ignore
        ws.append(["Update", "Current"])  # type: ignore

    return wb, ws, excel_path


def print_action_banner(action: str) -> None:
    """Print Grab/Down in bold to highlight detection."""
    colour = "\033[93m" if action == "Grab" else "\033[95m"  # yellow / magenta
    text = action.upper()
    print(f"\n{colour}\033[1m===== {text} DETECTED =====\033[0m\n")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #


def main() -> None:
    print("\033cStarting...\n")

    # Check data directory
    if not TEMPORARY_DATA_DIR.exists():
        sys.exit(f"Temporary data directory not found: {TEMPORARY_DATA_DIR}")

    # Prepare participant Excel
    wb, ws, excel_path = init_participant_excel(PARTICIPANT_DIR)
    print(f"Logging predictions to: {excel_path}")

    # Load reference header and models
    try:
        reference_columns = pd.read_csv(HEADER_CSV).drop(
            columns=["label"], errors="ignore"
        ).columns
    except FileNotFoundError:
        sys.exit(f"Header file not found: {HEADER_CSV}")

    try:
        scaler = joblib.load(SCALER_PATH)
        knn = joblib.load(KNN_PATH)
    except FileNotFoundError as err:
        sys.exit(f"Model or scaler file not found: {err}")

    # Wait for data folders
    try:
        while not any(p.is_dir() for p in TEMPORARY_DATA_DIR.iterdir()):
            print("Waiting for data, launch GetData.py")
            time.sleep(0.2)
            print(LINE_UP, end=LINE_CLEAR)
        print(f"\n{LINE_UP}", end=LINE_CLEAR)
    except KeyboardInterrupt:
        sys.exit("\nProgramme Stopped\n")

    print("\033c🟢 Real-time IMU-Only Prediction Started...")
    print("Mode: Pre-recorded dataset (offline)\n")

    current_vote = "Down"
    state_assurance: list[str] = []
    i = 0
    just_paused_for_action = False

    try:
        while True:
            folders = sorted(
                [
                    f.name
                    for f in TEMPORARY_DATA_DIR.iterdir()
                    if f.is_dir()
                ],
                key=numerical_sort,
            )

            recent_folders = folders[i : i + 3]

            if not recent_folders:
                print("No more samples found. Finishing.")
                break

            predictions: list[int] = []
            sample_used: list[str] = []

            for folder in recent_folders:
                folder_path = TEMPORARY_DATA_DIR / folder
                imu_file = folder_path / "imu.csv"

                if not imu_file.exists() or imu_file.stat().st_size == 0:
                    continue

                try:
                    imu_df = pd.read_csv(imu_file, header=None)
                    # Ensure this order matches imu.csv from GetData.py
                    imu_df.columns = [
                        "acc_x_1",
                        "acc_y_1",
                        "acc_z_1",
                        "gyro_x_1",
                        "gyro_y_1",
                        "gyro_z_1",
                        "acc_x_2",
                        "acc_y_2",
                        "acc_z_2",
                        "gyro_x_2",
                        "gyro_y_2",
                        "gyro_z_2",
                    ]
                except Exception as e:
                    print(f"⚠️ Skipping {folder}: {e}")
                    continue

                imu_feats = extract_imu_features(imu_df.iloc[:10])
                input_df = pd.DataFrame([imu_feats]).reindex(
                    columns=reference_columns, fill_value=0
                )
                scaled = scaler.transform(input_df)
                pred = knn.predict(scaled)[0]

                predictions.append(pred)
                sample_used.append(folder)

            if predictions:
                vote = Counter(predictions).most_common(1)[0][0]
                next_vote = LABEL_MAP[vote]

                print(
                    f"Prediction: {next_vote}\n"
                    f"📂 Samples Used: {', '.join(sample_used)}"
                )

                state_assurance.append(next_vote)
                if len(state_assurance) > 2 and state_assurance[-2] == state_assurance[-1]:
                    if next_vote != current_vote:
                        print(f"\n✨ Updated Prediction: {next_vote}")
                        update = next_vote

                        if next_vote in ["Grab", "Down"]:
                            print_action_banner(next_vote)
                            print("⏸ Pausing recognition for 6 seconds...\n")
                            just_paused_for_action = True
                    else:
                        update = ""

                    current_vote = next_vote
                else:
                    update = ""

                ws.append([update, next_vote])  # type: ignore
            else:
                print("⚠️ No valid samples found.")
                update = ""

            # Advance window in offline mode
            i += 1

            # Sleep handling
            if just_paused_for_action:
                time.sleep(6)
                just_paused_for_action = False
            else:
                time.sleep(1)

    except KeyboardInterrupt:
        print("\nKeyboard interrupt received. Stopping...")

    # Save Excel on exit
    if STOP_ALL:
        wb.save(excel_path)  # type: ignore
        print(f"Excel file has been saved successfully to: {excel_path}")

    print("\nProgramme Stopped\n")


if __name__ == "__main__":
    main()
