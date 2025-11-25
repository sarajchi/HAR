"""
Online IMU-only KNN prediction using streaming data collected by GetData.py.

Important:
    This script must be executed **at the same time** as:
        → 2_GetData.py
    which streams IMU data and writes Sample_* folders in Temporary_Data/.

    If GetData.py is not running, the system will keep waiting for data.

Features:
    - Always uses ONLINE MODE (last three samples).
    - Extracts time- and frequency-domain IMU features.
    - Applies trained scaler + KNN classifier.
    - Logs predictions into an automatically incremented Excel file.
    - Pauses for 6 seconds when detecting stable Grab / Down states.

This script is provided as companion code for the associated publication.
"""

if __name__ == "__main__":
    print("\033cStarting...\n")
    print("⚠️  Please ensure that '2_GetData.py' is running simultaneously.")
    print("    This script will wait until new IMU samples appear.\n")

import os
import sys
import time
import re
from pathlib import Path
from collections import Counter

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from scipy.fft import fft
from scipy.stats import skew, kurtosis, entropy

# =============================================================================
# Paths and Participant Excel Setup
# =============================================================================
time.sleep(5)
BASE_DIR = Path(__file__).resolve().parent
participant_path = BASE_DIR / "Participant_Data"
participant_path.mkdir(exist_ok=True)

# Create a unique Excel file
num = 1
excel_file = participant_path / f"participant_IMU_{num}.xlsx"
while excel_file.exists():
    num += 1
    excel_file = participant_path / f"participant_IMU_{num}.xlsx"

try:
    wb = load_workbook(excel_file)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Participant_IMU{num}"
    ws.append(["Update", "Current"])  # header

print(f"📄 Logging predictions to: {excel_file}")

# =============================================================================
# Configuration
# =============================================================================

label_map = {0: "Grab", 1: "Walk", 2: "Down"}

base_path = BASE_DIR / "Temporary_Data"

reference_columns = pd.read_csv("dataset_header.csv").drop(
    columns=["label"], errors="ignore"
).columns

scaler = joblib.load("scaler.pkl")
knn = joblib.load("knn_model.pkl")

LINE_UP = "\033[1A"
LINE_CLEAR = "\x1b[2K"
STOP_ALL = True

# =============================================================================
# Utility Functions
# =============================================================================

def numerical_sort(value: str):
    """Extract first integer from a string for natural sample sorting."""
    numbers = re.findall(r"\d+", value)
    return int(numbers[0]) if numbers else float("inf")


def extract_imu_features(df: pd.DataFrame) -> pd.Series:
    """Compute time-domain, statistical, and FFT features."""
    feats = {}
    for col in df.columns:
        sig = df[col].values

        feats[f"{col}_mean"] = np.mean(sig)
        feats[f"{col}_std"] = np.std(sig)
        feats[f"{col}_min"] = np.min(sig)
        feats[f"{col}_max"] = np.max(sig)
        feats[f"{col}_range"] = np.ptp(sig)
        feats[f"{col}_median"] = np.median(sig)
        feats[f"{col}_skew"] = skew(sig)
        feats[f"{col}_kurtosis"] = kurtosis(sig)
        feats[f"{col}_energy"] = np.sum(sig**2)
        feats[f"{col}_rms"] = np.sqrt(np.mean(sig**2))

        hist, _ = np.histogram(sig, bins=20, density=True)
        feats[f"{col}_entropy"] = entropy(hist + 1e-6)

        fft_vals = np.abs(fft(sig))[: len(sig) // 2]
        top_fft = np.sort(fft_vals)[-3:]
        for i, peak in enumerate(top_fft):
            feats[f"{col}_fft_peak_{i+1}"] = peak

    return pd.Series(feats)

# =============================================================================
# Wait for GetData.py to produce data
# =============================================================================

print("⏳ Waiting for incoming IMU data (from GetData.py)...\n")

try:
    while not any((base_path / f).is_dir() for f in os.listdir(base_path)):
        print("Waiting for Sample_* folders...", end="")
        time.sleep(0.3)
        print(LINE_UP, end=LINE_CLEAR)
except KeyboardInterrupt:
    sys.exit("\nProgramme Stopped\n")

print("✔️ IMU samples detected. Starting live prediction...\n")

# =============================================================================
# Online Prediction Loop
# =============================================================================

current_vote = "Down"
motor_state = "Down"
state_assurance = []

just_paused_for_action = False

print("\033c🟢 Real-time IMU-Only Prediction (ONLINE MODE)\n")

# initialise update so it always exists
update = ""

try:
    while True:
        # Load available sample folders
        folders = sorted(
            [f for f in os.listdir(base_path) if (base_path / f).is_dir()],
            key=numerical_sort
        )

        recent_folders = folders[-3:]  # ONLINE MODE

        predictions = []
        used_samples = []

        # ------------------------------------------------------------
        # Feature extraction from each new sample
        # ------------------------------------------------------------
        for folder in recent_folders:
            imu_path = base_path / folder / "imu.csv"

            if not imu_path.exists() or imu_path.stat().st_size == 0:
                continue

            try:
                imu_df = pd.read_csv(imu_path, header=None)
                imu_df.columns = [
                    "acc_x_1", "acc_y_1", "acc_z_1",
                    "gyro_x_1", "gyro_y_1", "gyro_z_1",
                    "acc_x_2", "acc_y_2", "acc_z_2",
                    "gyro_x_2", "gyro_y_2", "gyro_z_2"
                ]
            except Exception as e:
                print(f"⚠️ Skipping {folder}: {e}")
                continue

            feats = extract_imu_features(imu_df.iloc[:10])
            input_df = pd.DataFrame([feats]).reindex(columns=reference_columns, fill_value=0)
            scaled = scaler.transform(input_df)
            pred = knn.predict(scaled)[0]

            predictions.append(pred)
            used_samples.append(folder)

        # ------------------------------------------------------------
        # Decision logic
        # ------------------------------------------------------------
        if predictions:
            # reset update for this iteration; only set if state actually changes
            update = ""

            vote = Counter(predictions).most_common(1)[0][0]
            next_vote = label_map[vote]

            print(f"Prediction: {next_vote}\n📂 Using: {', '.join(used_samples)}")

            state_assurance.append(next_vote)

            if len(state_assurance) > 2 and state_assurance[-2] == state_assurance[-1]:
                if next_vote != current_vote:

                    print(f"\n✨ Updated Prediction: {next_vote}")
                    update = next_vote

                    # For Grab/Down, freeze the system for 6 seconds
                    if next_vote in ["Grab", "Down"]:
                        print("⏸ Freezing classification for 6 seconds...\n")
                        just_paused_for_action = True

                current_vote = next_vote

            # log both the last “update” (if any) and the current prediction
            ws.append([update, next_vote])

        else:
            print("⚠️ No valid IMU samples found.")
            update = ""

        # ------------------------------------------------------------
        # Timing control
        # ------------------------------------------------------------
        if just_paused_for_action:
            time.sleep(6)
            just_paused_for_action = False
        else:
            time.sleep(1)

except KeyboardInterrupt:
    pass

# =============================================================================
# Save Excel
# =============================================================================

if STOP_ALL:
    wb.save(excel_file)
    print("\n📄 Excel file saved successfully.")

print("\nProgramme Stopped\n")